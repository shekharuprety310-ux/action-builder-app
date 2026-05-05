#!/usr/bin/env python3
"""
Extract trade + product pricing from Deep Dive workbook into:
  - pricing_catalog.json (full catalog + cost-centre hints for the web app)
  - price_data.js (PRICE_DATA global for index.html)

Run from repo root:
  python3 build_price_data_from_deep_dive.py

Override workbook path:
  DEEP_DIVE_XLSM="path/to/file.xlsm" python3 build_price_data_from_deep_dive.py
"""
from __future__ import annotations

import json
import os
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
DEFAULT_XLSM = ROOT / "project files" / "Deep Dive Blank Template - Current.xlsm"
DOTX_JSON = ROOT / "cost_centres_dotx.json"

# Major "RATES" sections in Trade Price List -> key used in PRICE_DATA.trade (matches index.html CC_TO_TRADE_KEY)
RATES_TO_TRADE = {
    "BALUSTRADE RATES": "Balustrade",
    "BRICKLAYING RATES": "Brickwork",
    "PLASTERING RATES": "Plastering",
    "PAINTING RATES": "Painting & Rendering",
    "TILING RATES": "Tiling",
    "CONCRETING RATES": "Concreting",
    "ELECTRICAL RATES": "Electrical",
    "MAINS RATES": "Electrical Mains",
    "HEATING/COOLING RATES": "Heating",
    "PLUMBING RATES": "Plumbing",
    "DRAINAGE": "Plumbing",
    "PLUMBING": "Plumbing",
    "GAS": "Plumbing",
    "ALL INCLUSIVE RATES": "Plumbing",
    "ROOFING RATES": "Roofing",
    "SKYLIGHTS": "Roofing",
    # Sub-sections inherit last major; these are explicit if seen as majors
    "NEW HOME RATES": None,
    "RENOVATION RATES": None,
}

TIER_PREFIXES = ("Essentials", "Lifestyle", "Premium", "Prestige")


def _norm_money(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 4)
    return None


def compact_product(desc: str, price, code, supplier, uom, features, tier: str | None):
    p = _norm_money(price)
    if p is None:
        return None
    out = {"d": (desc or "").strip(), "p": p, "u": (uom or "ea").strip()}
    if supplier:
        out["s"] = str(supplier).strip()
    if code:
        out["c"] = str(code).strip()
    if features:
        out["f"] = str(features).strip()
    if tier:
        out["t"] = tier
    return out


def compact_trade(line_b: str, line_c: str, price, code, supplier, uom):
    p = _norm_money(price)
    if p is None:
        return None
    b = (line_b or "").strip()
    c = (line_c or "").strip()
    desc = f"{b} — {c}" if b and c else (b or c)
    out = {"d": desc.strip(), "p": p, "u": (uom or "ea").strip()}
    if supplier:
        out["s"] = str(supplier).strip()
    if code:
        out["c"] = str(code).strip()
    return out


def parse_trade_sheet(ws) -> dict[str, list]:
    rows = list(ws.iter_rows(values_only=True))
    trade_lists: dict[str, list] = defaultdict(list)
    major = "Trade Labouring"

    def is_data_row(r):
        if len(r) < 4:
            return False
        b, c, d = r[1], r[2], r[3]
        if b is None or not isinstance(b, str):
            return False
        if d is None or isinstance(d, str):
            return False
        return isinstance(d, (int, float))

    for r in rows[31:]:
        b = r[1] if len(r) > 1 else None
        if b and isinstance(b, str):
            u = b.strip().upper()
            if u in ("SERVICE DESCRIPTION",) or u.startswith("MAIN MENU") or u.startswith("PROJECT BUDGET"):
                continue
            if len(r) >= 4 and (r[2] is None or r[2] == "") and (r[3] is None or r[3] == ""):
                if "RATES" in u or u in ("DRAINAGE", "PLUMBING", "GAS", "SKYLIGHTS"):
                    key = RATES_TO_TRADE.get(u)
                    if key:
                        major = key
                    elif u not in ("NEW HOME RATES", "RENOVATION RATES"):
                        major = RATES_TO_TRADE.get(u, major)
                continue
        if is_data_row(r):
            code = r[4] if len(r) > 4 else None
            supplier = r[5] if len(r) > 5 else None
            uom = r[6] if len(r) > 6 else None
            item = compact_trade(r[1], r[2], r[3], code, supplier, uom)
            if item:
                trade_lists[major].append(item)

    return dict(trade_lists)


def _infer_tier(desc: str) -> str | None:
    if not desc:
        return None
    for t in TIER_PREFIXES:
        if desc.startswith(f"{t} Range") or desc.startswith(f"{t} range"):
            return t
    return None


def parse_products_sheet(ws) -> dict[str, list]:
    products: dict[str, list] = defaultdict(list)
    category = "GENERAL"

    for r in ws.iter_rows(min_row=6, values_only=True):
        b = r[1] if len(r) > 1 else None
        price = r[2] if len(r) > 2 else None
        if not b or not isinstance(b, str):
            continue
        b = b.strip()
        if b == "PRODUCT DESCRIPTION" or b.startswith("MAIN MENU") or b.startswith("PROJECT BUDGET"):
            continue
        # Category row: no numeric price in C
        if (price is None or price == "") and len(b) < 80:
            if b.isupper() or b in ("PLUMBING & SANITARY", "FLOOR COVERINGS") or "Custom:" in b:
                if "Custom:" not in b:
                    category = b.upper().replace("  ", " ")
                continue
        if isinstance(price, (int, float)):
            code = r[3] if len(r) > 3 else None
            supplier = r[4] if len(r) > 4 else None
            uom = r[5] if len(r) > 5 else None
            features = r[7] if len(r) > 7 else None
            tier = _infer_tier(b)
            item = compact_product(b, price, code, supplier, uom, features, tier)
            if item:
                products[category].append(item)

    # Normalize APPLIANCES parent — children are OVENS, COOKTOPS, ...
    out = {}
    for k, v in products.items():
        if k == "GENERAL" and not v:
            continue
        out[k] = [x for x in v if x and x.get("p") is not None]
    return out


def suggest_cost_centre_trade_map(cost_centre_names: list[str], trade_keys: list[str]) -> dict[str, str | None]:
    """Fuzzy map dotx cost centre labels to PRICE_DATA.trade keys."""
    hints: dict[str, str | None] = {}
    trade_keys_l = sorted(trade_keys, key=len, reverse=True)

    def norm(s: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", s.lower())

    n_trade = {norm(t): t for t in trade_keys}

    for cc in cost_centre_names:
        ccn = norm(cc)
        best = None
        for tk in trade_keys_l:
            tkn = norm(tk)
            if tkn and tkn in ccn:
                best = tk
                break
        if best is None:
            for tkn, orig in n_trade.items():
                if len(tkn) >= 4 and tkn in ccn:
                    best = orig
                    break
        hints[cc] = best
    return hints


def main():
    xlsm = Path(os.environ.get("DEEP_DIVE_XLSM", str(DEFAULT_XLSM))).expanduser()
    if not xlsm.is_file():
        raise SystemExit(f"Workbook not found: {xlsm}")

    wb = openpyxl.load_workbook(xlsm, read_only=True, data_only=True)
    trade = parse_trade_sheet(wb["Trade Price List"])
    products = parse_products_sheet(wb["Products Price List"])
    wb.close()

    cost_centre_names: list[str] = []
    if DOTX_JSON.is_file():
        with open(DOTX_JSON, encoding="utf-8") as f:
            dotx = json.load(f)
        cost_centre_names = list((dotx.get("cost_centres") or {}).keys())

    trade_map = suggest_cost_centre_trade_map(cost_centre_names, list(trade.keys()))

    catalog = {
        "meta": {
            "source": str(xlsm.relative_to(ROOT)) if xlsm.is_relative_to(ROOT) else str(xlsm),
            "trade_categories": list(trade.keys()),
            "product_categories": list(products.keys()),
        },
        "trade": trade,
        "products": products,
        "cost_centre_trade_map": trade_map,
    }

    out_json = ROOT / "pricing_catalog.json"
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(catalog, f, indent=2, ensure_ascii=False)

    # price_data.js — var PRICE_DATA for existing index.html
    js_path = ROOT / "price_data.js"
    js_body = (
        "// Auto-generated by build_price_data_from_deep_dive.py — re-run after Excel changes.\n"
        "var PRICE_DATA = "
        + json.dumps({"products": products, "trade": trade}, indent=2, ensure_ascii=False)
        + ";\n"
    )
    with open(js_path, "w", encoding="utf-8") as f:
        f.write(js_body)

    print("Wrote", out_json)
    print("Wrote", js_path)
    print("Trade buckets:", len(trade), "Product buckets:", len(products))


if __name__ == "__main__":
    main()
